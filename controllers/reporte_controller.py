from flask import Blueprint, request, jsonify, send_file
from firebase_admin import firestore
import pandas as pd
from io import BytesIO
from datetime import datetime

reporte_bp = Blueprint('reporte', __name__)
db = firestore.client()

# Convertir fechas y manejar errores
def parse_fecha(fecha_str):
    try:
        return datetime.fromisoformat(fecha_str)
    except Exception:
        raise ValueError("Formato de fecha inválido. Usa YYYY-MM-DD")

@reporte_bp.route('/obtener', methods=['GET'])
def obtener_reporte():
    try:
        fecha_inicio = parse_fecha(request.args.get('inicio'))
        fecha_fin = parse_fecha(request.args.get('fin'))

        query = db.collection("registros_asistencia")
        query = query.where(filter=firestore.FieldFilter('fecha_hora', '>=', fecha_inicio))
        query = query.where(filter=firestore.FieldFilter('fecha_hora', '<=', fecha_fin))

        asistencias = query.stream()

        datos = [doc.to_dict() for doc in asistencias]
        return jsonify(datos)

    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@reporte_bp.route('/excel', methods=['GET'])
def generar_excel():
    try:
        fecha_inicio = parse_fecha(request.args.get('inicio'))
        fecha_fin = parse_fecha(request.args.get('fin'))

        query = db.collection("registros_asistencia")
        query = query.where(filter=firestore.FieldFilter('fecha_hora', '>=', fecha_inicio))
        query = query.where(filter=firestore.FieldFilter('fecha_hora', '<=', fecha_fin))
        asistencias = query.stream()

        datos = [doc.to_dict() for doc in asistencias]
        if not datos:
            return jsonify({"error": "No hay datos en el rango especificado"}), 404

        # Normalizar datos
        for d in datos:
            d["nombre"] = d.get("nombre_usuario", "Desconocido")
            d["estado"] = d.get("estado", "")
            d["modo"] = d.get("modo", "")
            d["fecha"] = d.get("fecha_hora")

        df = pd.DataFrame(datos)
        df['fecha'] = pd.to_datetime(df['fecha']).dt.strftime('%d/%m/%Y %H:%M:%S')

        # Ordenar columnas
        columnas = {
            "nombre": "Nombre",
            "estado": "Estado",
            "modo": "Modo",
            "fecha": "Fecha y Hora"
        }
        df = df[list(columnas.keys())]
        df.rename(columns=columnas, inplace=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Asistencias')

            # Aplicar estilo con openpyxl
            workbook = writer.book
            sheet = writer.sheets['Asistencias']

            from openpyxl.styles import Font, Alignment, PatternFill

            # Estilo encabezado
            header_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col in sheet.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")

            # Ajuste de ancho de columnas automáticamente
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = length + 4

        output.seek(0)
        nombre = f"reporte_asistencia_{fecha_inicio.date()}_a_{fecha_fin.date()}.xlsx"
        return send_file(output, download_name=nombre, as_attachment=True)

    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500
